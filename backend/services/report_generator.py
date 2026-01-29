"""Report generator service for PDF and Excel exports"""
import io
from typing import Dict, Any, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

from backend.models.schemas import (
    MonthlyComparisonResult, ProductCostAnalysisResult,
    CostSimulationResult, BudgetComparisonResult,
    ReportType, ExportFormat
)


class ReportGenerator:
    """PDF 및 Excel 보고서 생성 서비스"""

    # 스타일 정의
    HEADER_FONT = Font(bold=True, size=12)
    TITLE_FONT = Font(bold=True, size=14)
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT_WHITE = Font(bold=True, color="FFFFFF")
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def generate_excel_report(
        self,
        data: Dict[str, Any],
        report_type: ReportType = ReportType.MONTHLY
    ) -> bytes:
        """Excel 보고서 생성"""
        wb = Workbook()

        # 요약 시트
        self._create_summary_sheet(wb, data)

        # 월간 분석 시트
        if 'monthly' in data:
            self._create_monthly_sheet(wb, data['monthly'])

        # 제품별 원가 시트
        if 'product_cost' in data:
            self._create_product_cost_sheet(wb, data['product_cost'])

        # 시뮬레이션 시트
        if 'simulation' in data:
            self._create_simulation_sheet(wb, data['simulation'])

        # 예산 비교 시트
        if 'budget' in data:
            self._create_budget_sheet(wb, data['budget'])

        # 기본 시트 삭제 (Sheet가 비어있으면)
        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb['Sheet']

        # 바이트로 반환
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _create_summary_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """요약 시트 생성"""
        ws = wb.create_sheet("요약", 0)

        # 제목
        ws['A1'] = "손익 분석 보고서"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')

        ws['A2'] = f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = Font(size=10, italic=True)

        row = 4

        # 월간 요약
        if 'monthly' in data:
            monthly: MonthlyComparisonResult = data['monthly']
            ws[f'A{row}'] = "■ 월간 손익 요약"
            ws[f'A{row}'].font = self.TITLE_FONT
            row += 1

            headers = ['구분', monthly.기준월, monthly.비교월, '변동액', '변동률']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.HEADER_FONT_WHITE
                cell.fill = self.HEADER_FILL
                cell.alignment = Alignment(horizontal='center')
            row += 1

            summary_items = [
                ('매출액', monthly.기준_요약.매출액, monthly.비교_요약.매출액),
                ('매출원가', monthly.기준_요약.매출원가, monthly.비교_요약.매출원가),
                ('매출총이익', monthly.기준_요약.매출총이익, monthly.비교_요약.매출총이익),
                ('판매관리비', monthly.기준_요약.판매관리비, monthly.비교_요약.판매관리비),
                ('영업이익', monthly.기준_요약.영업이익, monthly.비교_요약.영업이익),
            ]

            for name, prev, curr in summary_items:
                change_data = monthly.변동_요약.get(name, {})
                ws.cell(row=row, column=1, value=name)
                ws.cell(row=row, column=2, value=prev).number_format = '#,##0'
                ws.cell(row=row, column=3, value=curr).number_format = '#,##0'
                ws.cell(row=row, column=4, value=change_data.get('변동액', 0)).number_format = '#,##0'
                ws.cell(row=row, column=5, value=f"{change_data.get('변동률', 0):+.1f}%")

                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = self.THIN_BORDER
                row += 1

            row += 2

        # AI 코멘트
        if 'ai_comment' in data:
            ws[f'A{row}'] = "■ AI 분석 코멘트"
            ws[f'A{row}'].font = self.TITLE_FONT
            row += 1

            ai_comment = data['ai_comment']
            ws[f'A{row}'] = ai_comment

            # AI 코멘트 길이에 따라 동적으로 행 수 계산
            line_count = ai_comment.count('\n') + 1
            char_per_line = 80  # 대략적인 한 행당 문자 수
            estimated_lines = max(line_count, len(ai_comment) // char_per_line + 1)
            merge_rows = min(max(estimated_lines, 10), 30)  # 최소 10행, 최대 30행

            ws.merge_cells(f'A{row}:E{row + merge_rows - 1}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = 15 * merge_rows  # 행 높이 조정

        # 컬럼 너비 조정
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12

    def _create_monthly_sheet(self, wb: Workbook, monthly: MonthlyComparisonResult):
        """월간 분석 시트 생성"""
        ws = wb.create_sheet("월간분석")

        ws['A1'] = f"월간 손익 분석 ({monthly.기준월} vs {monthly.비교월})"
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:F1')

        # 상세 테이블
        row = 3
        headers = ['분류', '계정과목', monthly.기준월, monthly.비교월, '변동액', '변동률']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.HEADER_FONT_WHITE
            cell.fill = self.HEADER_FILL
            cell.border = self.THIN_BORDER
        row += 1

        for item in monthly.주요변동항목:
            ws.cell(row=row, column=1, value=item.분류).border = self.THIN_BORDER
            ws.cell(row=row, column=2, value=item.계정과목).border = self.THIN_BORDER
            ws.cell(row=row, column=3, value=item.기준금액).number_format = '#,##0'
            ws.cell(row=row, column=3).border = self.THIN_BORDER
            ws.cell(row=row, column=4, value=item.비교금액).number_format = '#,##0'
            ws.cell(row=row, column=4).border = self.THIN_BORDER
            ws.cell(row=row, column=5, value=item.변동액).number_format = '#,##0'
            ws.cell(row=row, column=5).border = self.THIN_BORDER
            ws.cell(row=row, column=6, value=f"{item.변동률:+.1f}%").border = self.THIN_BORDER

            # 변동률에 따른 색상
            if item.변동률 > 0:
                ws.cell(row=row, column=6).font = Font(color="FF0000")  # 빨강 (비용 증가)
            elif item.변동률 < 0:
                ws.cell(row=row, column=6).font = Font(color="0000FF")  # 파랑 (비용 감소)

            row += 1

        # 컬럼 너비
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_product_cost_sheet(self, wb: Workbook, product: ProductCostAnalysisResult):
        """제품별 원가 시트 생성"""
        ws = wb.create_sheet("제품별원가")

        ws['A1'] = f"제품별 원가 분석 ({product.기간})"
        ws['A1'].font = self.TITLE_FONT

        # 제품별 수익성 테이블
        row = 3
        headers = ['제품군', '매출액', '총원가', '매출총이익', '이익률']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.HEADER_FONT_WHITE
            cell.fill = self.HEADER_FILL
            cell.border = self.THIN_BORDER
        row += 1

        for p in product.제품별_분석:
            ws.cell(row=row, column=1, value=p.제품군).border = self.THIN_BORDER
            ws.cell(row=row, column=2, value=p.매출액).number_format = '#,##0'
            ws.cell(row=row, column=2).border = self.THIN_BORDER
            ws.cell(row=row, column=3, value=p.총원가).number_format = '#,##0'
            ws.cell(row=row, column=3).border = self.THIN_BORDER
            ws.cell(row=row, column=4, value=p.매출총이익).number_format = '#,##0'
            ws.cell(row=row, column=4).border = self.THIN_BORDER
            ws.cell(row=row, column=5, value=f"{p.매출총이익률:.1f}%").border = self.THIN_BORDER
            row += 1

        # 원가 구성비
        row += 2
        ws[f'A{row}'] = "■ 원가 구성비"
        ws[f'A{row}'].font = self.HEADER_FONT
        row += 1

        for category, ratio in product.원가구성비.items():
            ws.cell(row=row, column=1, value=category)
            ws.cell(row=row, column=2, value=f"{ratio}%")
            row += 1

    def _create_simulation_sheet(self, wb: Workbook, simulation: CostSimulationResult):
        """시뮬레이션 시트 생성"""
        ws = wb.create_sheet("시뮬레이션")

        ws['A1'] = "원가 변동 시뮬레이션 결과"
        ws['A1'].font = self.TITLE_FONT

        # 결과 요약
        row = 3
        ws.cell(row=row, column=1, value="구분")
        ws.cell(row=row, column=2, value="금액")
        ws.cell(row=row, column=1).font = self.HEADER_FONT
        ws.cell(row=row, column=2).font = self.HEADER_FONT
        row += 1

        results = [
            ('기준 매출원가', simulation.기준_매출원가),
            ('예상 매출원가', simulation.예상_매출원가),
            ('기준 영업이익', simulation.기준_영업이익),
            ('예상 영업이익', simulation.예상_영업이익),
            ('영업이익 변동액', simulation.영업이익_변동액),
        ]

        for name, value in results:
            ws.cell(row=row, column=1, value=name).border = self.THIN_BORDER
            ws.cell(row=row, column=2, value=value).number_format = '#,##0'
            ws.cell(row=row, column=2).border = self.THIN_BORDER
            row += 1

        ws.cell(row=row, column=1, value="영업이익 변동률").border = self.THIN_BORDER
        ws.cell(row=row, column=2, value=f"{simulation.영업이익_변동률:+.1f}%").border = self.THIN_BORDER

        # 항목별 영향
        row += 2
        ws[f'A{row}'] = "■ 원가 항목별 영향"
        ws[f'A{row}'].font = self.HEADER_FONT
        row += 1

        for 항목, 금액 in simulation.원가항목별_영향.items():
            if 금액 != 0:
                ws.cell(row=row, column=1, value=항목)
                cell = ws.cell(row=row, column=2, value=금액)
                cell.number_format = '#,##0'
                if 금액 > 0:
                    cell.font = Font(color="FF0000")
                else:
                    cell.font = Font(color="0000FF")
                row += 1

    def _create_budget_sheet(self, wb: Workbook, budget: BudgetComparisonResult):
        """예산 비교 시트 생성"""
        ws = wb.create_sheet("예산비교")

        ws['A1'] = f"예산 대비 실적 ({budget.기간})"
        ws['A1'].font = self.TITLE_FONT

        # 달성률 테이블
        row = 3
        headers = ['구분', '예산', '실적', '달성률']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.HEADER_FONT_WHITE
            cell.fill = self.HEADER_FILL
            cell.border = self.THIN_BORDER
        row += 1

        items = [
            ('매출액', budget.예산_요약.매출액, budget.실적_요약.매출액, budget.달성률['매출액']),
            ('매출원가', budget.예산_요약.매출원가, budget.실적_요약.매출원가, budget.달성률['매출원가']),
            ('영업이익', budget.예산_요약.영업이익, budget.실적_요약.영업이익, budget.달성률['영업이익']),
        ]

        for name, budget_val, actual_val, rate in items:
            ws.cell(row=row, column=1, value=name).border = self.THIN_BORDER
            ws.cell(row=row, column=2, value=budget_val).number_format = '#,##0'
            ws.cell(row=row, column=2).border = self.THIN_BORDER
            ws.cell(row=row, column=3, value=actual_val).number_format = '#,##0'
            ws.cell(row=row, column=3).border = self.THIN_BORDER
            ws.cell(row=row, column=4, value=f"{rate}%").border = self.THIN_BORDER

            # 달성률에 따른 색상
            if rate >= 100:
                ws.cell(row=row, column=4).font = Font(color="0000FF")
            else:
                ws.cell(row=row, column=4).font = Font(color="FF0000")
            row += 1

    def generate_pdf_report(
        self,
        data: Dict[str, Any],
        report_type: ReportType = ReportType.MONTHLY
    ) -> bytes:
        """PDF 보고서 생성 (WeasyPrint 사용)"""
        # HTML 템플릿 생성
        html_content = self._generate_html_report(data, report_type)

        try:
            from weasyprint import HTML
            pdf_bytes = HTML(string=html_content).write_pdf()
            return pdf_bytes
        except ImportError:
            # WeasyPrint가 설치되지 않은 경우
            raise ImportError("PDF 생성을 위해 weasyprint 패키지가 필요합니다.")

    def _generate_html_report(
        self,
        data: Dict[str, Any],
        report_type: ReportType
    ) -> str:
        """HTML 보고서 생성"""
        html = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        body { font-family: 'Malgun Gothic', sans-serif; margin: 40px; }
        h1 { color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; }
        h2 { color: #34495e; margin-top: 30px; }
        table { width: 100%; border-collapse: collapse; margin: 20px 0; }
        th { background-color: #3498db; color: white; padding: 10px; text-align: left; }
        td { padding: 8px; border: 1px solid #ddd; }
        tr:nth-child(even) { background-color: #f9f9f9; }
        .positive { color: #27ae60; }
        .negative { color: #e74c3c; }
        .summary-box { background: #ecf0f1; padding: 15px; border-radius: 5px; margin: 20px 0; }
        .ai-comment { background: #fff3cd; padding: 15px; border-radius: 5px; margin: 20px 0; }
    </style>
</head>
<body>
    <h1>손익 분석 보고서</h1>
    <p>생성일시: """ + datetime.now().strftime('%Y-%m-%d %H:%M') + """</p>
"""

        # 월간 분석 섹션
        if 'monthly' in data:
            monthly: MonthlyComparisonResult = data['monthly']
            html += f"""
    <h2>월간 손익 분석 ({monthly.기준월} → {monthly.비교월})</h2>
    <div class="summary-box">
        <p><strong>매출액:</strong> {monthly.비교_요약.매출액:,.0f}원
           <span class="{'positive' if monthly.변동_요약['매출액']['변동률'] > 0 else 'negative'}">
           ({monthly.변동_요약['매출액']['변동률']:+.1f}%)</span></p>
        <p><strong>영업이익:</strong> {monthly.비교_요약.영업이익:,.0f}원
           <span class="{'positive' if monthly.변동_요약['영업이익']['변동률'] > 0 else 'negative'}">
           ({monthly.변동_요약['영업이익']['변동률']:+.1f}%)</span></p>
    </div>
"""

        # AI 코멘트
        if 'ai_comment' in data:
            html += f"""
    <h2>AI 분석 코멘트</h2>
    <div class="ai-comment">
        {data['ai_comment'].replace(chr(10), '<br>')}
    </div>
"""

        html += """
</body>
</html>
"""
        return html


# 싱글톤 인스턴스
report_generator = ReportGenerator()
