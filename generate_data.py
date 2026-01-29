# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

np.random.seed(42)

months = [f'2024년 {m}월' for m in range(1, 13)]
months_short = [f'{m}월' for m in range(1, 13)]

# 계절 지수
건재_계절 = [0.85, 0.90, 1.05, 1.15, 1.20, 1.10, 0.95, 0.90, 1.10, 1.15, 1.00, 0.80]
가전_계절 = [0.90, 0.85, 0.95, 1.00, 1.10, 1.25, 1.30, 1.20, 1.00, 0.95, 1.05, 1.15]
가스_계절 = [1.3, 1.2, 1.0, 0.9, 0.8, 0.8, 0.8, 0.8, 0.9, 1.0, 1.2, 1.4]

# 기준값 (월평균, 억원)
건재_기준 = 28.5
가전_기준 = 14.2
기타_기준 = 2.3

def generate_pnl_data(months_list, add_noise=True):
    rows = []

    def add_row(분류, 계정, values):
        row = {'분류': 분류, '계정과목': 계정}
        for i, m in enumerate(months_list):
            row[m] = int(values[i])
        rows.append(row)

    def noise(x):
        if add_noise:
            return x * (1 + np.random.uniform(-0.03, 0.03))
        return x

    # 매출
    건재_values = [int(noise(건재_기준 * 건재_계절[i] * 1e8)) for i in range(12)]
    가전_values = [int(noise(가전_기준 * 가전_계절[i] * 1e8)) for i in range(12)]
    기타_values = [int(noise(기타_기준 * 1e8)) for i in range(12)]
    총매출_values = [건재_values[i] + 가전_values[i] + 기타_values[i] for i in range(12)]

    add_row('매출', '제품매출-건재용', 건재_values)
    add_row('매출', '제품매출-가전용', 가전_values)
    add_row('매출', '제품매출-산업용기타', 기타_values)

    # 원재료비
    add_row('매출원가', '원재료비-냉연강판(POSCO)', [int(noise(총매출_values[i] * 0.42)) for i in range(12)])
    add_row('매출원가', '원재료비-도료(KCC,삼화)', [int(noise(총매출_values[i] * 0.065)) for i in range(12)])
    add_row('매출원가', '원재료비-아연도금재', [int(noise(총매출_values[i] * 0.035)) for i in range(12)])
    add_row('매출원가', '원재료비-화학약품', [int(noise(총매출_values[i] * 0.025)) for i in range(12)])
    add_row('매출원가', '원재료비-포장재', [int(noise(총매출_values[i] * 0.015)) for i in range(12)])

    # 노무비
    add_row('매출원가', '노무비-생산직급여', [int(noise(3.2e8)) for i in range(12)])
    add_row('매출원가', '노무비-생산직상여', [int(0.85e8) if i in [0, 6, 11] else 0 for i in range(12)])
    add_row('매출원가', '노무비-품질관리팀', [int(noise(0.78e8)) for i in range(12)])
    add_row('매출원가', '노무비-생산관리팀', [int(noise(0.65e8)) for i in range(12)])

    # 제조경비
    add_row('매출원가', '제조경비-전력비', [int(noise(1.6e8 * (1 + 0.1 * (건재_계절[i] + 가전_계절[i] - 2)))) for i in range(12)])
    add_row('매출원가', '제조경비-가스비(LNG)', [int(noise(0.9e8 * 가스_계절[i])) for i in range(12)])
    add_row('매출원가', '제조경비-감가상각비', [124000000 for i in range(12)])
    add_row('매출원가', '제조경비-수선유지비', [int(noise(0.48e8)) for i in range(12)])
    add_row('매출원가', '제조경비-외주가공비', [int(noise(총매출_values[i] * 0.015)) for i in range(12)])
    add_row('매출원가', '제조경비-소모품비', [int(noise(0.25e8)) for i in range(12)])
    add_row('매출원가', '제조경비-보험료', [18000000 for i in range(12)])

    # 판매관리비
    add_row('판매관리비', '인건비-급여', [int(noise(1.98e8)) for i in range(12)])
    add_row('판매관리비', '인건비-퇴직급여', [int(noise(0.33e8)) for i in range(12)])
    add_row('판매관리비', '인건비-복리후생비', [int(noise(0.42e8)) for i in range(12)])
    add_row('판매관리비', '물류비-운반비', [int(noise(총매출_values[i] * 0.042)) for i in range(12)])
    add_row('판매관리비', '물류비-포장비', [int(noise(총매출_values[i] * 0.010)) for i in range(12)])
    add_row('판매관리비', '판매비-광고선전비', [int(noise(0.15e8)) for i in range(12)])
    add_row('판매관리비', '판매비-접대비', [int(noise(0.095e8)) for i in range(12)])
    add_row('판매관리비', '일반관리비-여비교통비', [int(noise(0.12e8)) for i in range(12)])
    add_row('판매관리비', '일반관리비-통신비', [int(noise(0.055e8)) for i in range(12)])
    add_row('판매관리비', '일반관리비-세금과공과', [int(noise(0.22e8)) for i in range(12)])
    add_row('판매관리비', '일반관리비-감가상각비', [34000000 for i in range(12)])
    add_row('판매관리비', '일반관리비-지급수수료', [int(noise(0.28e8)) for i in range(12)])
    add_row('판매관리비', '일반관리비-대손상각비', [int(총매출_values[i] * 0.002) for i in range(12)])

    # 영업외손익
    add_row('영업외손익', '영업외수익-이자수익', [int(noise(0.032e8)) for i in range(12)])
    add_row('영업외손익', '영업외비용-이자비용', [int(noise(0.28e8)) for i in range(12)])
    add_row('영업외손익', '영업외수익-외환차익', [int(0.05e8 * np.random.uniform(0.5, 2.0)) for i in range(12)])
    add_row('영업외손익', '영업외비용-외환차손', [int(0.04e8 * np.random.uniform(0.5, 2.0)) for i in range(12)])
    add_row('영업외손익', '영업외수익-잡이익', [int(noise(0.015e8)) for i in range(12)])
    add_row('영업외손익', '영업외비용-잡손실', [int(noise(0.008e8)) for i in range(12)])

    return rows

def save_excel(rows, filename, title, header_color):
    wb = Workbook()
    ws = wb.active
    ws.title = "데이터"

    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 제목
    ws.merge_cells('A1:N1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # 헤더
    months_list = list(rows[0].keys())[2:]  # 분류, 계정과목 제외
    headers = ['분류', '계정과목'] + months_list
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # 데이터
    for r, row_data in enumerate(rows, 4):
        ws.cell(row=r, column=1, value=row_data['분류']).border = thin_border
        ws.cell(row=r, column=2, value=row_data['계정과목']).border = thin_border
        for c, m in enumerate(months_list, 3):
            cell = ws.cell(row=r, column=c, value=row_data[m])
            cell.number_format = '#,##0'
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right')

    # 컬럼 너비
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 28
    for i in range(3, 15):
        ws.column_dimensions[chr(64+i)].width = 14

    wb.save(filename)

def print_summary(rows, months_list, label):
    매출합계 = sum(sum(row[m] for m in months_list) for row in rows if row['분류'] == '매출')
    원가합계 = sum(sum(row[m] for m in months_list) for row in rows if row['분류'] == '매출원가')
    판관비합계 = sum(sum(row[m] for m in months_list) for row in rows if row['분류'] == '판매관리비')

    print(f"\n{'='*50}")
    print(f"{label}")
    print('='*50)
    print(f"매출액:     {매출합계/1e8:>10,.0f}억원")
    print(f"매출원가:   {원가합계/1e8:>10,.0f}억원 (원가율 {원가합계/매출합계*100:.1f}%)")
    print(f"매출총이익: {(매출합계-원가합계)/1e8:>10,.0f}억원 (총이익률 {(매출합계-원가합계)/매출합계*100:.1f}%)")
    print(f"판관비:     {판관비합계/1e8:>10,.0f}억원")
    print(f"영업이익:   {(매출합계-원가합계-판관비합계)/1e8:>10,.0f}억원 (영업이익률 {(매출합계-원가합계-판관비합계)/매출합계*100:.1f}%)")

# 1. 손익계산서 (실적) 생성
print("데이터 생성 중...")
pnl_rows = generate_pnl_data(months, add_noise=True)
save_excel(pnl_rows, 'data/sample/손익계산서_2024년.xlsx', '㈜한국컬러강판 월별 손익계산서 (2024년 실적)', '1F4E79')
print_summary(pnl_rows, months, '2024년 손익계산서 (실적) 요약')

# 2. 예산 생성 (실적보다 목표치 약간 높게)
np.random.seed(100)  # 다른 시드
건재_기준 = 29.0  # 예산은 목표치
가전_기준 = 14.5
기타_기준 = 2.4

budget_rows = generate_pnl_data(months_short, add_noise=False)
save_excel(budget_rows, 'data/sample/예산_2024년.xlsx', '㈜한국컬러강판 2024년 예산', '2E7D32')
print_summary(budget_rows, months_short, '2024년 예산 요약')

print(f"\n{'='*50}")
print("파일 생성 완료!")
print('='*50)
print("1. data/sample/손익계산서_2024년.xlsx")
print("2. data/sample/예산_2024년.xlsx")
